#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
seaice.io.icxl.py : function to import import ice core data from xlsx spreadsheet
"""

__name__ = "icxl"
__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "1.1"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "moggier@alaska.edu"
__status__ = "dev"
__date__ = "2017/09/13"
__comment__ = "loadxl.py contained function to import ice core data from xlsx spreadsheet"
__CoreVersion__ = 1.1

import datetime
import logging
import os

import dateutil
import numpy as np
import openpyxl
import pandas as pd
import seaice

__all__ = ["import_ic_path", "import_ic_list", "import_ic_sourcefile", "list_ic", "list_ic_path", "make_ic_sourcefile"]

TOL =1e-6
subvariable_dict = {'conductivity': ['conductivity measurement temperature']}

variable_2_sheet = {'temperature': 'T_ice',
                    'salinity': 'S_ice',
                    'conductivity': 'S_ice',
                    'specific conductance': 'S_ice',
                    'd18O': 'S_ice',
                    'dD': 'S_ice',
                    'Vf_oil': 'Vf_oil', 'oil volume fraction': 'Vf_oil',  # MOSIDEO project
                    'Wf_oil': 'Wf_oil', 'oil weight fraction': 'Vf_oil',  # MOSIDEO project
                    'oil content': 'oil_content',  # CMI project
                    'oil mass': 'Vf_oil', 'm_oil': 'Vf_oil'
                    # 'seawater': 'seawater',
                    # 'sediment': 'sediment',
                    # 'Chla': 'algal_pigment',
                    # 'chlorophyl a': 'algal_pigment',
                    # 'Phae': 'algal_pigment'
                    }


def import_ic_path(ic_path, variables=None, v_ref='top', drop_empty=False):
    """
    :param ic_path:
        string, path to the xlsx ice core spreadsheet
    :param variables:
        list of string, variables to import. If not defined, all variable will be imported.
    :param v_ref:
        'top' or 'bottom', vertical reference. top for ice/snow or ice/air surface, bottom for ice/water interface
    :return:
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])

    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_name = wb.sheetnames
    ws_summary = wb['summary']  # load the data from the summary sheet

    name = ws_summary['C21'].value

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)

    # convert ice core spreadsheet to last version
    if version < __CoreVersion__:
        update_spreadsheet(ic_path, v_ref=v_ref)
        logger.info("Updating ice core spreadsheet %s to last version (%s)" % (name, str(__CoreVersion__)))
        wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
        ws_name = wb.sheetnames
        ws_summary = wb['summary']  # load the data from the summary sheet
        version = ws_summary['C3'].value

    n_row_collection = 22
    logger.info("importing data for %s" % name)

    if isinstance(ws_summary['C2'].value, datetime.datetime):
        if isinstance(ws_summary['D2'].value, datetime.time):
            date = datetime.datetime.combine(ws_summary['C2'].value, ws_summary['D2'].value)
            if ws_summary['E2'].value is not None and dateutil.tz.gettz(ws_summary['E2'].value):
                tz = dateutil.tz.gettz(ws_summary['E2'].value)
                date = date.replace(tzinfo=tz)
            else:
                logger.info("\t(%s) timezone unavailable." % name)
        else:
            date = ws_summary['C2'].value
            if ws_summary['D2'].value is not None and dateutil.tz.gettz(ws_summary['D2'].value):
                tz = dateutil.tz.gettz(ws_summary['D2'].value)
                date = date.replace(tzinfo=tz)
            else:
                logger.info("\t(%s) timezone unavailable." % name)
    else:
        logger.warning("\t(%s) date unavailable" % name)
        date = None

    origin = ws_summary['C5'].value

    if isinstance(ws_summary['C6'].value, (float, int)) or isinstance(ws_summary['D6'], (float, int)):
        lat = ws_summary['C6'].value
        lon = ws_summary['D6'].value
    elif ws_summary['C6'].value and ws_summary['D6'].value:
        logger.info("\t(%s) lat/lon not defined in decimal degree" % name)
        lat = np.nan
        lon = np.nan
    else:
        logger.info("\t(%s) lat/lon unknown" % name)
        lat = np.nan
        lon = np.nan

    if isinstance(ws_summary['C9'].value, (float, int)):
        snow_depth = np.array([ws_summary['C9'].value]).astype(float)
        n_snow = 1
        while ws_summary.cell(row=9, column=3+n_snow).value is not None:
            snow_depth = np.concatenate((snow_depth, np.array([ws_summary.cell(row=9, column=3+n_snow).value])))
            n_snow +=1
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')
    else:
        snow_depth = np.array([np.nan])

    if isinstance(ws_summary['C10'].value, (float, int)):
        freeboard = np.array([ws_summary['C10'].value])
        n_temp = 1
        while ws_summary.cell(row=10, column=3+n_temp).value is not None:
            if isinstance(ws_summary.cell(row=10, column=3+n_temp).value, (float, int)):
                freeboard = np.concatenate((freeboard, np.array([ws_summary.cell(row=10, column=3 + n_snow).value])))
            else:
                logger.info("(%s)\tfreeboard cell %s not a float" % (name, openpyxl.utils.get_column_letter(3 + n_temp)+str(9)))
            n_temp += 1
        freeboard = pd.to_numeric(freeboard, errors='coerce')
    else:
        freeboard = np.array([np.nan])

    if isinstance(ws_summary['C11'].value, (float, int)):
        ice_thickness = np.array([ws_summary['C11'].value])
        n_temp = 1
        while ws_summary.cell(row=11, column=3+n_temp).value:
            if isinstance(ws_summary.cell(row=11, column=3+n_temp).value, (float, int)):
                ice_thickness = np.concatenate((ice_thickness, np.array([ws_summary.cell(row=11, column=3+n_snow).value])))
            else:
                logger.info("\t(%s) ice_thickness cell %s not a float" % (name, openpyxl.utils.get_column_letter(3+n_temp)+str(9)))
            n_temp += 1
        ice_thickness = pd.to_numeric(ice_thickness, errors='coerce')
    else:
        ice_thickness = np.array([np.nan])

    core = seaice.Core(name, date, origin, lat, lon, ice_thickness, freeboard, snow_depth)

    # temperature
    if ws_summary['C15'].value:
        core.t_air = ws_summary['C15'].value
    if ws_summary['C16'].value:
        core.t_snow_surface = ws_summary['C16'].value
    if ws_summary['C17'].value:
        core.t_ice_surface = ws_summary['C17'].value
    if ws_summary['C18'].value:
        core.t_water = ws_summary['C18'].value

    # sampling protocol
    m_col = 3
    if ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value is not None:
        core.protocol = ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value
    else:
        core.protocol = 'N/A'

    # core collection
    while (ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)] is not None and
           ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value is not None):
        core.add_to_collection(ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value)
        m_col += 1

    # comment
    if ws_summary['C33'].value is not None:
        core.add_comment(ws_summary['C33'].value)

    # import all variables
    if variables is None:
        sheets = [sheet for sheet in ws_name if (sheet not in ['summary', 'abreviation', 'locations', 'lists', 'Vf_oil_calculation']) and
                     (sheet.lower().find('fig') == -1)]
        for sheet in sheets:
            ws_variable = wb[sheet]
            profile = read_profile(ws_variable, variables=None, version=version, v_ref=v_ref)

            if drop_empty:
                profile.drop_empty_property()

            if not profile.empty:
                if profile.get_name() is not core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                 % (ic_path, core.name, profile.get_name()))
                else:
                    core.add_profile(profile)
                    logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.get_property())))
            else:
                logger.info('(%s) no data to import from %s ' % (core.name, sheet))
    else:
        if not isinstance(variables, list):
            if variables.lower().find('state variable')+1:
                variables = ['temperature', 'salinity']
            else:
                variables = [variables]

        _imported_variables = []
        for variable in variables:
            if variable_2_sheet[variable] in ws_name and variable not in _imported_variables:
                sheet = variable_2_sheet[variable]
                ws_variable = wb[sheet]

                variable2import = [var for var in variables if var in inverse_dict(variable_2_sheet)[sheet]]

                profile = read_profile(ws_variable, variables=variable2import, version=version, v_ref=v_ref)

                if profile.get_name() is not core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                 % (ic_path, core.name, profile.name()))
                elif not profile.empty:
                    core.add_profile(profile)
                    logger.info(' (%s) data imported with success: %s' % (profile.get_name(), profile.get_property()))
                else:
                    _temp = [variable for variable in profile.get_variable() if profile[variable].isnull().all()]
                    if _temp.__len__() > 1:
                        logger.info(' (%s) no data to import: %s ' % (profile.get_name(), ", ".join(_temp)))
                    else:
                        logger.info('(%s) no variable to import' % name)

                _imported_variables +=variable2import

    return core


def import_ic_list(ic_list, variables=None, v_ref='top', verbose=False, drop_empty=False):
    """
    :param ic_list:
            array, array contains absolute filepath for the cores
    :param variables:
    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)

    ic_dict = {}
    inexisting_ic_list = []
    for ic_path in ic_list:
        if verbose:
            print('Importing data from %s' % ic_path)
        if not os.path.exists(ic_path):
            logger.warning("%s does not exists in core directory" % ic_path.split('/')[-1])
            inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
        else:
            ic_data = import_ic_path(ic_path, variables=variables, v_ref=v_ref, drop_empty=drop_empty)
            if not ic_data.variables():
                inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
                logger.warning("%s have no properties profile" % (ic_data.name))
            else:
                ic_dict[ic_data.name] = ic_data

    logging.info("Import ice core lists completed")
    if inexisting_ic_list.__len__()>0:
        logger.info("%s core does not exits. Removing from collection" % ', '.join(inexisting_ic_list))

    for ic in inexisting_ic_list:
        for ic2 in ic_dict.keys():
            if ic in ic_dict[ic2].collection:
                ic_dict[ic2].del_from_collection(ic)
                logger.info("remove %s from %s collection" % (ic, ic2))
    return ic_dict


def import_ic_sourcefile(f_path, variables=None, ic_dir=None, v_ref='top', drop_empty=False):
    """
    :param filepath:
            string, absolute path to the file containing either the absolute path of the cores (1 path by line) or the
            core names (1 core by line). In this last case if core_dir is None core_dir is the directory contianing the
            file.
    :param variables:

    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)
    logger.info('Import ice core from source file: %s' % f_path)

    if ic_dir is not None:
        with open(f_path) as f:
            ics = sorted([os.path.join(ic_dir, line.strip()) for line in f if not line.strip().startswith('#')])
    else:
        with open(f_path) as f:
            ics = sorted([line.strip() for line in f if not line.strip().startswith('#')])

    print(ics)

    return import_ic_list(ics, variables=variables, v_ref=v_ref, drop_empty=drop_empty)


# read profile
def read_profile(ws_variable, variables=None, version=__CoreVersion__, v_ref='top', fill_missing=False):
    """
    :param ws_variable:
        openpyxl.worksheet
    :param variables:
    :param version:
    :param v_ref:
        top, or bottom
    :param fill_missing:

    """
    logger = logging.getLogger(__name__)

    if version == 1:
        row_data_start = 6
    elif version == 1.1:
        row_data_start = 8
        if ws_variable['C4'].value:
            v_ref = ws_variable['C4'].value
    else:
        logger.error("ice core spreadsheet version not defined")

    sheet_2_data = {'S_ice': [row_data_start, 'ABC', 'DEFG', 'J'],
                    'T_ice': [row_data_start, 'A', 'B', 'C'],
                    'Vf_oil': [row_data_start, 'ABC', 'DEFG', 'H']}
    # TODO: add other sheets for seawater, sediment, CHla, Phae, stratigraphy

    #                'stratigraphy': [row_data_start, 'AB', 'C', 'D'],
    #                'seawater': [row_data_start, 'A', 'DEFGF', 'G']}

    name = ws_variable['C1'].value

    # define section
    headers = ['y_low', 'y_mid', 'y_sup']
    # Continuous profile
    if not ws_variable.title in sheet_2_data:
        profile = seaice.core.profile.Profile()
    else:
        if sheet_2_data[ws_variable.title][1].__len__() == 1:
            y_mid = np.array([ws_variable[sheet_2_data[ws_variable.title][1] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row + 1)]).astype(float)
            y_low = np.nan * np.ones(y_mid.__len__())
            y_sup = np.nan * np.ones(y_mid.__len__())
        # Step profile
        elif sheet_2_data[ws_variable.title][1].__len__() == 2:
            y_low = np.array([ws_variable[sheet_2_data[ws_variable.title][1][0] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row+1)]).astype(float)
            y_sup = np.array([ws_variable[sheet_2_data[ws_variable.title][1][1] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row+1)]).astype(float)
            y_mid = (y_low + y_sup) / 2
            y_mid = y_mid.astype(float)
        else:
            y_low = np.array([ws_variable[sheet_2_data[ws_variable.title][1][0] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row+1)]).astype(float)
            y_sup = np.array([ws_variable[sheet_2_data[ws_variable.title][1][1] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row+1)]).astype(float)
            y_mid = np.array([ws_variable[sheet_2_data[ws_variable.title][1][2] + str(row)].value
                              for row in range(sheet_2_data[ws_variable.title][0], ws_variable.max_row+1)]).astype(float)

            # if y_mid is not defined, y_mid = (y_low+y_sup)/2
            if np.isnan(y_mid).any():
                if (np.isnan(y_mid).any() or np.isnan(y_mid).any()):
                    y_mid = (y_low + y_sup) / 2
                    logger.info(
                        '\t(%s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
                        % ws_variable.title)
                else:
                    logger.warning(
                        '\t(%s) not all y_mid exists, some element of y_low or y_sup are not defined. Unable to'
                        'compute y_mid = (y_low+y_sup)/2' % ws_variable.title)

        data = np.array([y_low, y_mid, y_sup])

        # read data
        min_col = sheet_2_data[ws_variable.title][2][0]
        min_col = openpyxl.utils.column_index_from_string(min_col)
        max_col = sheet_2_data[ws_variable.title][2][-1]
        max_col = openpyxl.utils.column_index_from_string(max_col)
        min_row = sheet_2_data[ws_variable.title][0]
        max_row = min_row + y_mid.__len__()-1

        _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
                 for row in ws_variable.iter_cols(min_col, max_col, min_row, max_row)]
        data = np.vstack([data, np.array(_data).astype(float)])
        data = data.transpose()
        header_offset = 3
        variable_headers = [ws_variable[col + str(row_data_start - header_offset)].value
                            for col in sheet_2_data[ws_variable.title][2]]
        headers += variable_headers

        # add comment to dataframe
        headers_float = headers.copy()
        headers += ['comment']
        comment = [ws_variable[sheet_2_data[ws_variable.title][3][0] + str(row)].value
                   for row in range(min_row, max_row + 1)]
        data = np.hstack([data, np.atleast_2d(comment).transpose()])

        # fill missing section with np.nan
        if fill_missing:
            idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]

            for ii_idx in idx:
                empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
                empty += [np.nan] * (variable_headers.__len__()+1)
            data = np.vstack([data, empty])

        # assemble profile dataframe
        profile = pd.DataFrame(data, columns=headers)
        profile = profile.sort_values(by='y_low')

        # ice thickness
        try:
            length = float(ws_variable['C2'].value)
        except:
            logger.info('(%s) no ice core length' % name)
            length = np.nan
        else:
            if length == 'n/a':
                logger.info('(%s) ice core length is not available (n/a)' % name)
                length = np.nan
            elif not isinstance(length, (int, float)):
                logger.info('%s ice core length is not a number' % name)
                length = np.nan
        profile['length'] = length
        headers_float += ['length']

        # clean profile if there is no depth entry
        profile = profile.dropna(axis=0, subset=['y_low', 'y_mid', 'y_sup'])

        # set vertical references
        profile['v_ref'] = v_ref

        # set ice core name for profile
        profile['name'] = name

        # remove subvariables from variables
        for variable in variable_headers:
            if variable in subvariable_dict:
                for subvariable in subvariable_dict[variable]:
                    variable_headers.remove(subvariable)
        profile['variable'] = ', '.join(variable_headers)

        # set depth and property type to float
        profile[headers_float] = profile[headers_float].astype(float)

        profile = seaice.core.profile.Profile(profile)

        # remove variable not in variables
        if variables is not None:
            for property in profile.get_property():
                if property not in variables:
                    profile.delete_property(variable)
    return profile


# create list or source
def list_ic(dirpath, fileext):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return ic_list: list
        list of ice core path
    """
    logger = logging.getLogger(__name__)

    ics_set = set([f for f in os.listdir(dirpath) if f.endswith(fileext)])
    logger.info("Found %i ice core datafile in %s" % (ics_set.__len__(), dirpath))
    return ics_set


def list_ic_path(dirpath, fileext):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return ic_list: list
        list of ice core path
    """
    logger = logging.getLogger(__name__)

    ics_set = list_ic(dirpath=dirpath, fileext=fileext)
    ic_paths_set = set([os.path.join(os.path.realpath(dirpath), f) for f in ics_set])
    return ic_paths_set


def make_ic_sourcefile(dirpath, fileext, source_filepath=None):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return source_file: str
        filepath to the text file containing ice core filepath with absolute path.
    """
    logger = logging.getLogger(__name__)

    ic_paths_set = list_ic_path(dirpath, fileext)

    if source_filepath is None:
        source_filepath = os.path.join(os.path.realpath(dirpath), 'ic_list.txt')

    with open(source_filepath, 'w') as f:
        for ic_path in ic_paths_set:
            f.write(ic_path + "\n")

    return source_filepath


# updater
def update_spreadsheet(ic_path, v_ref='top', backup=True):
    """
    update_spreadsheet update an ice core file to the latest ice core file version (__CoreVersion__).

    :param ic_path: str
        Path to the ice core file to update
    :param v_ref: int, 'top' or 'bottom', Default 'top'
        Set the zero vertical reference of the core. If 'top, zero vertical reference is set at the ice/snow or ice/air
        interface. If 'bottom, zero vertical reference is set at the ice/water interface
    :param verbose: int
        verbosity of the function.
    :param backup: bool, default True
        If True, backup the previous file version in the subdirectory 'folder-COREVERSION'
    :return nothing:

    USAGE:
        update_spreadhseet(ic_path)
    """
    logger = logging.getLogger(__name__)

    import shutil
    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0
    else:
        logger.info("\t ice core file path %s" % ic_path)

    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_summary = wb['summary']  # load the data from the summary sheet

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
        if version < __CoreVersion__:
            logger.info("Updating core data to latest version %.1f" % __CoreVersion__)
    else:
        logger.error("\t%s ice core fiel version not unavailable" % ic_path)

    # backup old version
    if backup:
        backup_dir = os.path.join(os.path.dirname(ic_path), 'version-'+str(version))
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        shutil.copyfile(ic_path, os.path.join(backup_dir, os.path.basename(ic_path)))

    while version < __CoreVersion__:
        # update from 1.0 to 1.1
        if version == 1:
            version = 1.1
            ws_summary['C3'] = version
            ws_summary = delete_row(ws_summary, 22)

            # loop through variables
            if "S_ice" in wb.sheetnames:
                ws = wb["S_ice"]
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                row_start = 4
                ws = delete_column(ws, 'E', row_start, ws.max_row)
                ws['D5'] = 'salinity'
                ws['D6'] = 'S'
                ws = move_column(ws, 'E', 'J', row_start, ws.max_row)
                ws['E5'] = 'conductivity'
                ws['E6'] = 'σ'
                ws = move_column(ws, 'F', 'K', row_start, ws.max_row)
                ws['F5'] = 'conductivity measurement temperature'
                ws['F6'] = 'T_σ'
                ws = move_column(ws, 'G', 'J', row_start, ws.max_row)
                ws['G5'] = 'specific conductance'
                ws['G6'] = 'κ'
                delete_column(ws, 'K', row_start, ws.max_row)

            if "T_ice" in wb.sheetnames:
                ws = wb["T_ice"]
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd'
                ws['B6'] = 'T'

            if "oil_content" in wb.sheetnames:
                ws = wb["oil_content"]
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd_1'
                ws['B6'] = 'd_2'
                ws['C6'] = 'd'
                ws['D6'] = 'V_i'
                ws['E6'] = 'h_menisc'
                ws['F6'] = 'd_menisc'
                ws['G6'] = 'd_center'
        wb.save(ic_path)


def add_row(ws, row_number):
    """
    :param ws:
    :param row_number:
    :return:
    """
    max_row = ws.max_row
    for row in range(row_number, ws.max_row + 1):
        new_row = row_number + max_row + 1 - row
        old_row = row_number + max_row - row
        for col in range(1, ws.max_column+1):
            ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
    for col in range(1, ws.max_column+1):
        ws.cell(row=row_number, column=col).value = ""
    return ws


def delete_row(ws, row_number):
    """

    :param ws:
    :param row_number:
    :return:
    """
    max_row = ws.max_row
    for row in range(row_number, max_row):
        new_row = row
        old_row = row+1
        for col in range(1, ws.max_column+1):
            ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
    for col in range(1, ws.max_column+1):
        ws.cell(row=max_row+1, column=col).value = ""
    return ws


def delete_column(ws, target_col, start_row=None, end_row=None):
    """
    :param ws:
    :param target_col:
    :param start_row:
    :param end_row:
    :return:
    """
    if start_row is None:
        start_row = ws.min_row
    if end_row is None:
        end_row = ws.max_row
    if not isinstance(target_col, int):
        target_col = openpyxl.utils.column_index_from_string(target_col)

    max_col = ws.max_column
    if np.alltrue([ws.cell(row=row, column=ws.max_column).value is None for row in range(start_row, ws.max_row)]):
        max_col = max_col - 1

    if not target_col == max_col:
        for col in range(target_col, max_col):
            new_col = col
            old_col = col + 1
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=max_col).value = ""

    return ws


def move_column(ws, target_col, source_col, start_row=None, end_row=None):
    """
    :param ws:
    :param target_col:
    :param source_col:
    :param start_row:
    :param end_row:
    :return:
    """

    if start_row is None:
        start_row = ws.min_row
    if end_row is None:
        end_row = ws.max_row
    if not isinstance(target_col, int):
        target_col = openpyxl.utils.column_index_from_string(target_col)
    if not isinstance(source_col, int):
        source_col = openpyxl.utils.column_index_from_string(source_col)

    max_col = ws.max_column
    if np.alltrue([ws.cell(row=row, column=ws.max_column).value is None for row in range(start_row, ws.max_row)]):
        max_col = max_col - 1

    # insert column in target column
    for col in range(target_col, max_col+1):
        new_col = target_col + max_col - col + 1
        old_col = target_col + max_col - col
        # print(openpyxl.utils.get_column_letter(old_col), openpyxl.utils.get_column_letter(new_col))
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value

    # copy source col to target column
    if target_col < source_col:
        source_col = source_col + 1
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=target_col).value = ws.cell(row=row, column=source_col).value

    ws = delete_column(ws, source_col, start_row, end_row)

    return ws


def print_column(ws, col):
    """
    :param ws:
    :param col:
    :return:
    """
    if ~isinstance(col, int):
        col = openpyxl.utils.column_index_from_string(col)
    for row in range(1, ws.max_row + 1):
        print(row, ws.cell(row=row, column=col).value)


def print_row(ws, row):
    """
    :param ws:
    :param row:
    :return:
    """
    col_string = ''
    for col in range(1, ws.max_column + 1):
        col_string += openpyxl.utils.get_column_letter(col) + ':' + str(
            ws.cell(row=row, column=col).value) + '\t'
    print(col_string)


def inverse_dict(map):
    """
    return the inverse of a dictionnary with non-unique values
    :param map: dictionnary
    :return inv_map: dictionnary
    """
    revdict = {}
    for k, v in map.items():
        revdict.setdefault(v, []).append(k)

    return revdict